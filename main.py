import random
import threading
import tkinter as tk
import webbrowser
from datetime import datetime
from pathlib import Path
from tkinter import filedialog, messagebox, simpledialog, ttk
from urllib.parse import quote

from ocr_utils import extrair_pilotos_da_area_transferencia, extrair_pilotos_de_arquivo
from sorteio_core import (
    aplicar_correcoes_nomes,
    formatar_mensagem_whatsapp,
    mensagem_karts_minimos,
    mesclar_atribuicoes,
    parse_csv_conteudo,
    realizar_sorteio,
    validar_repescagem,
    validar_sorteio,
)
from storage_desktop import carregar, salvar

try:
    from openpyxl import load_workbook

    TEM_EXCEL = True
except ImportError:
    TEM_EXCEL = False

try:
    from PIL import Image, ImageDraw, ImageFont

    TEM_PIL = True
except ImportError:
    TEM_PIL = False


class SorteioKartApp:
    def __init__(self, root: tk.Tk) -> None:
        self.root = root
        self.root.title("Sorteio de Kart")
        self.root.geometry("980x720")
        self.root.minsize(900, 650)

        self.dados = carregar()
        self.atribuicoes: list[tuple[str, str]] = []
        self.ultimo_seed: int | None = None
        self.tema_escuro = self.dados.get("tema", "dark") == "dark"
        self._salvando = False

        self._build_ui()
        self._carregar_evento_ativo()
        self._aplicar_tema()

    def _build_ui(self) -> None:
        self.main = ttk.Frame(self.root, padding=16)
        self.main.pack(fill=tk.BOTH, expand=True)

        topo = ttk.Frame(self.main)
        topo.pack(fill=tk.X, pady=(0, 12))

        ttk.Label(topo, text="Sorteio de Pilotos x Karts", font=("Segoe UI", 18, "bold")).pack(
            side=tk.LEFT
        )

        ttk.Button(topo, text="Tema", command=self._alternar_tema).pack(side=tk.RIGHT, padx=4)
        ttk.Button(topo, text="Histórico", command=self._abrir_historico).pack(side=tk.RIGHT, padx=4)

        evento_frame = ttk.LabelFrame(self.main, text="Evento", padding=10)
        evento_frame.pack(fill=tk.X, pady=(0, 12))

        self.combo_eventos = ttk.Combobox(evento_frame, state="readonly", width=40)
        self.combo_eventos.pack(side=tk.LEFT, fill=tk.X, expand=True)
        self.combo_eventos.bind("<<ComboboxSelected>>", self._trocar_evento)

        ttk.Button(evento_frame, text="Novo", command=self._novo_evento).pack(side=tk.LEFT, padx=6)
        ttk.Button(evento_frame, text="Excluir", command=self._excluir_evento).pack(side=tk.LEFT)

        colunas = ttk.Frame(self.main)
        colunas.pack(fill=tk.BOTH, expand=True)

        self._build_pilotos_panel(colunas)
        self._build_karts_panel(colunas)

        excluidos_frame = ttk.LabelFrame(self.main, text="Karts indisponíveis", padding=10)
        excluidos_frame.pack(fill=tk.X, pady=12)
        self.karts_excluidos_text = tk.Text(excluidos_frame, height=3, font=("Segoe UI", 11))
        self.karts_excluidos_text.pack(fill=tk.X)
        self.karts_excluidos_text.bind("<<Modified>>", self._on_texto_modificado)
        self.karts_excluidos_text.bind("<KeyRelease>", self._on_texto_modificado)

        acoes = ttk.Frame(self.main)
        acoes.pack(fill=tk.X, pady=8)
        ttk.Button(acoes, text="Sortear", command=self._sortear).pack(side=tk.LEFT)
        self.btn_repescar = ttk.Button(acoes, text="Repescar", command=self._repescar, state=tk.DISABLED)
        self.btn_repescar.pack(side=tk.LEFT, padx=8)
        ttk.Button(acoes, text="Limpar tudo", command=self._limpar).pack(side=tk.LEFT, padx=8)

        resultado = ttk.LabelFrame(self.main, text="Resultado do sorteio", padding=12)
        resultado.pack(fill=tk.BOTH, expand=True)

        resultado_acoes = ttk.Frame(resultado)
        resultado_acoes.pack(fill=tk.X, pady=(0, 8))
        self.lbl_seed = ttk.Label(resultado_acoes, text="", foreground="gray")
        self.lbl_seed.pack(side=tk.LEFT)

        self.btn_whatsapp = ttk.Button(
            resultado_acoes, text="WhatsApp", command=self._exportar_whatsapp, state=tk.DISABLED
        )
        self.btn_whatsapp.pack(side=tk.RIGHT, padx=4)
        self.btn_imagem = ttk.Button(
            resultado_acoes, text="Exportar imagem", command=self._exportar_imagem, state=tk.DISABLED
        )
        self.btn_imagem.pack(side=tk.RIGHT, padx=4)
        self.btn_apresentacao = ttk.Button(
            resultado_acoes, text="Apresentação", command=self._abrir_apresentacao, state=tk.DISABLED
        )
        self.btn_apresentacao.pack(side=tk.RIGHT, padx=4)

        self.result_tree = ttk.Treeview(resultado, columns=("piloto", "kart"), show="headings", height=8)
        self.result_tree.heading("piloto", text="Piloto")
        self.result_tree.heading("kart", text="Kart")
        self.result_tree.column("piloto", width=350, anchor=tk.W)
        self.result_tree.column("kart", width=120, anchor=tk.CENTER)
        scroll = ttk.Scrollbar(resultado, orient=tk.VERTICAL, command=self.result_tree.yview)
        self.result_tree.configure(yscrollcommand=scroll.set)
        self.result_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scroll.pack(side=tk.RIGHT, fill=tk.Y)

        self.status = ttk.Label(self.main, text="Pronto para sortear. Dados salvos automaticamente.")
        self.status.pack(anchor=tk.W, pady=(8, 0))

    def _build_pilotos_panel(self, parent: ttk.Frame) -> None:
        frame = ttk.LabelFrame(parent, text="Pilotos", padding=12)
        frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(0, 8))

        acoes = ttk.Frame(frame)
        acoes.pack(fill=tk.X, pady=(0, 8))
        ttk.Button(acoes, text="Ler área de transferência", command=self._importar_clipboard).pack(
            side=tk.LEFT
        )
        ttk.Button(acoes, text="Ler arquivo", command=self._importar_arquivo).pack(side=tk.LEFT, padx=6)
        ttk.Button(acoes, text="Colar lista", command=self._colar_pilotos).pack(side=tk.LEFT, padx=6)
        ttk.Button(acoes, text="CSV", command=self._importar_csv).pack(side=tk.LEFT, padx=6)
        if TEM_EXCEL:
            ttk.Button(acoes, text="Excel", command=self._importar_excel).pack(side=tk.LEFT, padx=6)

        self.pilotos_text = tk.Text(frame, height=12, font=("Segoe UI", 11))
        self.pilotos_text.pack(fill=tk.BOTH, expand=True)
        self.pilotos_text.bind("<<Modified>>", self._on_texto_modificado)
        self.pilotos_text.bind("<KeyRelease>", self._on_texto_modificado)

    def _build_karts_panel(self, parent: ttk.Frame) -> None:
        frame = ttk.LabelFrame(parent, text="Karts", padding=12)
        frame.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True, padx=(8, 0))
        self.karts_text = tk.Text(frame, height=12, font=("Segoe UI", 11))
        self.karts_text.pack(fill=tk.BOTH, expand=True)
        self.karts_text.bind("<<Modified>>", self._on_texto_modificado)
        self.karts_text.bind("<KeyRelease>", self._on_texto_modificado)
        ttk.Label(frame, text="Um kart por linha, sem repetir.", foreground="gray").pack(anchor=tk.W, pady=(6, 0))

    def _evento_ativo(self) -> dict:
        return self.dados["eventos"][self.dados["evento_ativo"]]

    def _ler_linhas_widget(self, widget: tk.Text) -> list[str]:
        return [l.strip() for l in widget.get("1.0", tk.END).splitlines() if l.strip()]

    def _ler_formulario(self) -> dict:
        return {
            "pilotos": self._ler_linhas_widget(self.pilotos_text),
            "karts": self._ler_linhas_widget(self.karts_text),
            "karts_excluidos": self._ler_linhas_widget(self.karts_excluidos_text),
        }

    def _on_texto_modificado(self, _event=None) -> None:
        if self._salvando:
            return
        self.root.after(400, self._salvar_automatico)

    def _salvar_automatico(self) -> None:
        if self._salvando:
            return
        form = self._ler_formulario()
        evento = self._evento_ativo()
        evento["pilotos"] = form["pilotos"]
        evento["karts"] = form["karts"]
        evento["karts_excluidos"] = form["karts_excluidos"]
        salvar(self.dados)

    def _atualizar_combo_eventos(self) -> None:
        nomes = [(e["nome"], eid) for eid, e in self.dados["eventos"].items()]
        self.combo_eventos["values"] = [n for n, _ in nomes]
        ativo = self._evento_ativo()
        self.combo_eventos.set(ativo["nome"])

    def _carregar_evento_ativo(self) -> None:
        self._salvando = True
        evento = self._evento_ativo()
        self.pilotos_text.delete("1.0", tk.END)
        self.pilotos_text.insert("1.0", "\n".join(evento.get("pilotos", [])))
        self.karts_text.delete("1.0", tk.END)
        self.karts_text.insert("1.0", "\n".join(evento.get("karts", [])))
        self.karts_excluidos_text.delete("1.0", tk.END)
        self.karts_excluidos_text.insert("1.0", "\n".join(evento.get("karts_excluidos", [])))
        self._salvando = False
        self._atualizar_combo_eventos()

    def _novo_evento(self) -> None:
        nome = simpledialog.askstring("Novo evento", "Nome do evento:", parent=self.root)
        if not nome or not nome.strip():
            return
        self._salvar_automatico()
        import uuid

        eid = str(uuid.uuid4())
        self.dados["eventos"][eid] = {
            "nome": nome.strip(),
            "pilotos": [],
            "karts": [],
            "karts_excluidos": [],
        }
        self.dados["evento_ativo"] = eid
        salvar(self.dados)
        self._carregar_evento_ativo()
        self._limpar_resultado()
        self.status.config(text=f'Evento "{nome.strip()}" criado.')

    def _excluir_evento(self) -> None:
        if len(self.dados["eventos"]) <= 1:
            messagebox.showwarning("Atenção", "É necessário manter pelo menos um evento.")
            return
        if not messagebox.askyesno("Confirmar", "Excluir este evento?"):
            return
        eid = self.dados["evento_ativo"]
        del self.dados["eventos"][eid]
        self.dados["evento_ativo"] = next(iter(self.dados["eventos"]))
        salvar(self.dados)
        self._carregar_evento_ativo()
        self._limpar_resultado()

    def _trocar_evento(self, _event=None) -> None:
        self._salvar_automatico()
        nome = self.combo_eventos.get()
        for eid, ev in self.dados["eventos"].items():
            if ev["nome"] == nome:
                self.dados["evento_ativo"] = eid
                break
        salvar(self.dados)
        self._carregar_evento_ativo()
        self._limpar_resultado()

    def _aplicar_pilotos(self, linhas: list[str], substituir: bool) -> None:
        corrigidos = aplicar_correcoes_nomes(linhas)
        if substituir:
            self.pilotos_text.delete("1.0", tk.END)
            self.pilotos_text.insert("1.0", "\n".join(corrigidos) + "\n")
        else:
            atuais = self._ler_linhas_widget(self.pilotos_text)
            self.pilotos_text.delete("1.0", tk.END)
            self.pilotos_text.insert("1.0", "\n".join(atuais + corrigidos) + "\n")
        self._salvar_automatico()
        total = len(self._ler_linhas_widget(self.pilotos_text))
        msg = f"{len(corrigidos)} piloto(s) importado(s).\n\n{mensagem_karts_minimos(total)}"
        messagebox.showinfo("Pilotos importados", msg)
        self.status.config(text=mensagem_karts_minimos(total))

    def _preview_pilotos(self, pilotos: list[str], titulo: str = "Confirmar pilotos") -> None:
        janela = tk.Toplevel(self.root)
        janela.title(titulo)
        janela.geometry("480x420")
        janela.transient(self.root)
        janela.grab_set()

        ttk.Label(
            janela,
            text=f"Foram encontrados {len(pilotos)} piloto(s). Revise e confirme.",
            wraplength=440,
        ).pack(anchor=tk.W, padx=16, pady=(16, 8))

        texto = tk.Text(janela, height=14, font=("Segoe UI", 11))
        texto.pack(fill=tk.BOTH, expand=True, padx=16, pady=(0, 8))
        texto.insert("1.0", "\n".join(pilotos))

        modo = tk.StringVar(value="substituir")
        opcoes = ttk.Frame(janela)
        opcoes.pack(fill=tk.X, padx=16)
        ttk.Radiobutton(opcoes, text="Substituir lista", variable=modo, value="substituir").pack(anchor=tk.W)
        ttk.Radiobutton(opcoes, text="Adicionar à lista", variable=modo, value="adicionar").pack(anchor=tk.W)

        botoes = ttk.Frame(janela)
        botoes.pack(fill=tk.X, padx=16, pady=16)

        def confirmar() -> None:
            linhas = [l.strip() for l in texto.get("1.0", tk.END).splitlines() if l.strip()]
            if not linhas:
                messagebox.showwarning("Atenção", "Informe pelo menos um piloto.", parent=janela)
                return
            self._aplicar_pilotos(linhas, modo.get() == "substituir")
            janela.destroy()

        ttk.Button(botoes, text="Confirmar", command=confirmar).pack(side=tk.RIGHT)
        ttk.Button(botoes, text="Cancelar", command=janela.destroy).pack(side=tk.RIGHT, padx=8)

    def _colar_pilotos(self) -> None:
        janela = tk.Toplevel(self.root)
        janela.title("Colar lista")
        janela.geometry("480x360")
        janela.transient(self.root)
        janela.grab_set()
        ttk.Label(janela, text="Cole os nomes (um por linha):").pack(anchor=tk.W, padx=16, pady=12)
        texto = tk.Text(janela, height=12, font=("Segoe UI", 11))
        texto.pack(fill=tk.BOTH, expand=True, padx=16)
        modo = tk.StringVar(value="substituir")
        ttk.Radiobutton(janela, text="Substituir", variable=modo, value="substituir").pack(anchor=tk.W, padx=16)
        ttk.Radiobutton(janela, text="Adicionar", variable=modo, value="adicionar").pack(anchor=tk.W, padx=16)

        def confirmar() -> None:
            linhas = [l.strip() for l in texto.get("1.0", tk.END).splitlines() if l.strip()]
            if not linhas:
                messagebox.showwarning("Atenção", "Cole pelo menos um piloto.", parent=janela)
                return
            self._aplicar_pilotos(linhas, modo.get() == "substituir")
            janela.destroy()

        ttk.Button(janela, text="Confirmar", command=confirmar).pack(pady=12)

    def _importar_csv(self) -> None:
        caminho = filedialog.askopenfilename(filetypes=[("CSV", "*.csv *.txt"), ("Todos", "*.*")])
        if not caminho:
            return
        conteudo = Path(caminho).read_text(encoding="utf-8", errors="ignore")
        itens = parse_csv_conteudo(conteudo)
        if not itens:
            messagebox.showwarning("Atenção", "Nenhum piloto encontrado.")
            return
        self._preview_pilotos(itens, "Importar CSV")

    def _importar_excel(self) -> None:
        if not TEM_EXCEL:
            messagebox.showinfo("Excel", "Instale openpyxl: pip install openpyxl")
            return
        caminho = filedialog.askopenfilename(filetypes=[("Excel", "*.xlsx *.xls")])
        if not caminho:
            return
        wb = load_workbook(caminho, read_only=True)
        folha = wb.active
        itens = aplicar_correcoes_nomes(
            [str(row[0]).strip() for row in folha.iter_rows(values_only=True) if row and row[0]]
        )
        if not itens:
            messagebox.showwarning("Atenção", "Nenhum piloto encontrado.")
            return
        self._preview_pilotos(itens, "Importar Excel")

    def _importar_arquivo(self) -> None:
        caminho = filedialog.askopenfilename(
            filetypes=[("Imagens", "*.png *.jpg *.jpeg *.bmp *.webp"), ("Todos", "*.*")]
        )
        if not caminho:
            return
        self._executar_ocr("Lendo imagem...", lambda: extrair_pilotos_de_arquivo(caminho))

    def _importar_clipboard(self) -> None:
        self._executar_ocr("Lendo print...", extrair_pilotos_da_area_transferencia)

    def _executar_ocr(self, msg: str, func) -> None:
        self.status.config(text=msg)
        self.root.config(cursor="watch")

        def tarefa() -> None:
            try:
                pilotos = func()
                self.root.after(0, lambda: self._finalizar_ocr(pilotos))
            except Exception as erro:
                self.root.after(0, lambda: self._falha_ocr(str(erro)))

        threading.Thread(target=tarefa, daemon=True).start()

    def _falha_ocr(self, msg: str) -> None:
        self.root.config(cursor="")
        messagebox.showerror("Erro OCR", msg)

    def _finalizar_ocr(self, pilotos: list[str]) -> None:
        self.root.config(cursor="")
        if not pilotos:
            messagebox.showwarning("OCR", "Nenhum piloto identificado.")
            return
        self._preview_pilotos(pilotos)

    def _registrar_historico(self, tipo: str) -> None:
        evento = self._evento_ativo()
        self.dados.setdefault("historico", []).insert(
            0,
            {
                "evento_nome": evento["nome"],
                "tipo": tipo,
                "seed": self.ultimo_seed,
                "atribuicoes": self.atribuicoes,
                "data": datetime.now().isoformat(),
            },
        )
        self.dados["historico"] = self.dados["historico"][:50]
        salvar(self.dados)

    def _exibir_resultado(self, atribuicoes: list[tuple[str, str]], seed: int) -> None:
        self.atribuicoes = atribuicoes
        self.ultimo_seed = seed
        for item in self.result_tree.get_children():
            self.result_tree.delete(item)
        for piloto, kart in atribuicoes:
            self.result_tree.insert("", tk.END, values=(piloto, kart))
        self.lbl_seed.config(text=f"Código de auditoria: {seed}")
        for btn in (self.btn_whatsapp, self.btn_imagem, self.btn_apresentacao, self.btn_repescar):
            btn.config(state=tk.NORMAL)

    def _limpar_resultado(self) -> None:
        self.atribuicoes = []
        self.ultimo_seed = None
        for item in self.result_tree.get_children():
            self.result_tree.delete(item)
        self.lbl_seed.config(text="")
        for btn in (self.btn_whatsapp, self.btn_imagem, self.btn_apresentacao, self.btn_repescar):
            btn.config(state=tk.DISABLED)

    def _sortear(self) -> None:
        form = self._ler_formulario()
        ok, erro, disponiveis = validar_sorteio(form["pilotos"], form["karts"], form["karts_excluidos"])
        if not ok:
            messagebox.showerror("Erro", erro)
            return
        atribuicoes, seed = realizar_sorteio(form["pilotos"], disponiveis)
        self._exibir_resultado(atribuicoes, seed)
        self._registrar_historico("sorteio")
        sobra = len(disponiveis) - len(form["pilotos"])
        self.status.config(
            text=f"Sorteio concluído. Código: {seed}."
            + (f" {sobra} kart(s) sem uso." if sobra > 0 else "")
        )

    def _repescar(self) -> None:
        if not self.atribuicoes:
            return
        janela = tk.Toplevel(self.root)
        janela.title("Repescagem")
        janela.geometry("400x360")
        janela.transient(self.root)
        janela.grab_set()

        vars_check: dict[str, tk.BooleanVar] = {}
        for piloto, kart in self.atribuicoes:
            var = tk.BooleanVar()
            vars_check[piloto] = var
            ttk.Checkbutton(janela, text=f"{piloto} (Kart {kart})", variable=var).pack(anchor=tk.W, padx=16, pady=2)

        def confirmar() -> None:
            selecionados = [p for p, v in vars_check.items() if v.get()]
            if not selecionados:
                messagebox.showwarning("Atenção", "Selecione pelo menos um piloto.", parent=janela)
                return
            form = self._ler_formulario()
            ok, erro, disponiveis = validar_repescagem(
                form["pilotos"],
                form["karts"],
                form["karts_excluidos"],
                self.atribuicoes,
                selecionados,
            )
            if not ok:
                messagebox.showerror("Erro", erro, parent=janela)
                return
            novas, seed = realizar_sorteio(selecionados, disponiveis)
            mescladas = mesclar_atribuicoes(self.atribuicoes, novas)
            self._exibir_resultado(mescladas, seed)
            self._registrar_historico("repescagem")
            self.status.config(text=f"Repescagem concluída. Novo código: {seed}.")
            janela.destroy()

        ttk.Button(janela, text="Repescar", command=confirmar).pack(pady=12)

    def _limpar(self) -> None:
        if not messagebox.askyesno("Confirmar", "Limpar listas e resultado?"):
            return
        self.pilotos_text.delete("1.0", tk.END)
        self.karts_text.delete("1.0", tk.END)
        self.karts_excluidos_text.delete("1.0", tk.END)
        self._limpar_resultado()
        self._salvar_automatico()
        self.status.config(text="Listas e resultado limpos.")

    def _exportar_whatsapp(self) -> None:
        if not self.atribuicoes:
            return
        evento = self._evento_ativo()
        msg = formatar_mensagem_whatsapp(self.atribuicoes, self.ultimo_seed, evento["nome"])
        webbrowser.open(f"https://wa.me/?text={quote(msg)}")

    def _exportar_imagem(self) -> None:
        if not self.atribuicoes or not TEM_PIL:
            messagebox.showinfo("Imagem", "Requer Pillow instalado.")
            return
        caminho = filedialog.asksaveasfilename(
            defaultextension=".png", filetypes=[("PNG", "*.png")], initialfile="sorteio-kart.png"
        )
        if not caminho:
            return
        altura = 120 + len(self.atribuicoes) * 48 + 40
        img = Image.new("RGB", (720, altura), "#1a1a2e")
        draw = ImageDraw.Draw(img)
        draw.rectangle([0, 0, 720, 6], fill="#e94560")
        draw.text((32, 32), "Sorteio de Kart", fill="white")
        evento = self._evento_ativo()
        draw.text((32, 68), f"{evento['nome']} — Código {self.ultimo_seed}", fill="#a0a0b8")
        y = 120
        for i, (piloto, kart) in enumerate(self.atribuicoes):
            if i % 2 == 0:
                draw.rectangle([24, y, 696, y + 48], fill="#24243e")
            draw.text((40, y + 12), piloto, fill="white")
            draw.text((620, y + 12), str(kart), fill="#4ade80")
            y += 48
        img.save(caminho)
        messagebox.showinfo("Imagem", f"Salvo em:\n{caminho}")

    def _abrir_apresentacao(self) -> None:
        if not self.atribuicoes:
            return
        janela = tk.Toplevel(self.root)
        janela.title("Apresentação")
        janela.attributes("-fullscreen", True)
        janela.configure(bg="#0a0a14")

        evento = self._evento_ativo()
        tk.Label(
            janela, text=evento["nome"], font=("Segoe UI", 28, "bold"), bg="#0a0a14", fg="white"
        ).pack(pady=(40, 8))
        tk.Label(
            janela,
            text=f"Código {self.ultimo_seed}",
            font=("Segoe UI", 14),
            bg="#0a0a14",
            fg="#a0a0b8",
        ).pack()

        frame = tk.Frame(janela, bg="#0a0a14")
        frame.pack(fill=tk.BOTH, expand=True, padx=40, pady=20)
        for piloto, kart in self.atribuicoes:
            linha = tk.Frame(frame, bg="#0a0a14")
            linha.pack(fill=tk.X, pady=8)
            tk.Label(linha, text=piloto, font=("Segoe UI", 22), bg="#0a0a14", fg="white").pack(
                side=tk.LEFT
            )
            tk.Label(linha, text=str(kart), font=("Segoe UI", 28, "bold"), bg="#0a0a14", fg="#4ade80").pack(
                side=tk.RIGHT
            )
        ttk.Button(janela, text="Fechar", command=janela.destroy).pack(pady=20)

    def _abrir_historico(self) -> None:
        janela = tk.Toplevel(self.root)
        janela.title("Histórico")
        janela.geometry("520x400")
        texto = tk.Text(janela, font=("Segoe UI", 10))
        texto.pack(fill=tk.BOTH, expand=True, padx=12, pady=12)
        for item in self.dados.get("historico", []):
            data = datetime.fromisoformat(item["data"]).strftime("%d/%m/%Y %H:%M")
            resumo = ", ".join(f"{p}→{k}" for p, k in item.get("atribuicoes", [])[:4])
            texto.insert(
                tk.END,
                f"{item.get('evento_nome')} | {data} | {item.get('tipo')} | Código {item.get('seed')}\n{resumo}\n\n",
            )
        texto.config(state=tk.DISABLED)

    def _alternar_tema(self) -> None:
        self.tema_escuro = not self.tema_escuro
        self.dados["tema"] = "dark" if self.tema_escuro else "light"
        salvar(self.dados)
        self._aplicar_tema()

    def _aplicar_tema(self) -> None:
        if self.tema_escuro:
            bg, fg, entry_bg = "#1a1a2e", "#f5f5f7", "#24243e"
        else:
            bg, fg, entry_bg = "#f3f4f8", "#1a1a2e", "#ffffff"
        self.root.configure(bg=bg)
        for widget in (self.pilotos_text, self.karts_text, self.karts_excluidos_text):
            widget.configure(bg=entry_bg, fg=fg, insertbackground=fg)


def main() -> None:
    root = tk.Tk()
    SorteioKartApp(root)
    root.mainloop()


if __name__ == "__main__":
    main()
